"""
app.py — Aplikasi Streamlit: Training Dataset Generator
Memproses PDF → Daftar Isi, Keywords, Ringkasan (Excel) + JSONL
"""

import io
import json
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine import engine1_daftar_isi, engine2_keywords, engine3_ringkasan, engine4_jsonl


# ─────────────────────────────────────────────
# Konfigurasi halaman
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Training Dataset Generator",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ─────────────────────────────────────────────
# CSS styling
# ─────────────────────────────────────────────

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 2rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        text-align: center;
    }
    .main-header h1 { color: #e94560; margin: 0; font-size: 2.2rem; }
    .main-header p { color: #a8b2d8; margin: 0.5rem 0 0; font-size: 1rem; }

    .engine-card {
        background: #1e1e2e;
        border: 1px solid #333;
        border-radius: 10px;
        padding: 1.2rem;
        margin-bottom: 1rem;
    }
    .engine-card h4 { color: #cdd6f4; margin: 0 0 0.5rem; }
    .engine-card p  { color: #7f849c; font-size: 0.85rem; margin: 0; }

    .status-box {
        padding: 0.7rem 1rem;
        border-radius: 8px;
        margin: 0.4rem 0;
        font-size: 0.9rem;
    }
    .status-success { background: #1e3a2f; color: #a6e3a1; border-left: 3px solid #a6e3a1; }
    .status-pending  { background: #2a2a3e; color: #cba6f7; border-left: 3px solid #cba6f7; }
    .status-error    { background: #3a1e1e; color: #f38ba8; border-left: 3px solid #f38ba8; }

    .result-container {
        background: #1e1e2e;
        border: 1px solid #444;
        border-radius: 10px;
        padding: 1.2rem;
        margin-top: 1rem;
    }
    .download-section {
        background: #16213e;
        border-radius: 12px;
        padding: 1.5rem;
        margin-top: 1.5rem;
        border: 1px solid #0f3460;
    }
    .stButton > button {
        background: linear-gradient(135deg, #e94560, #c62a47);
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 600;
        width: 100%;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #c62a47, #a01f39);
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# Inisialisasi session state / cache
# ─────────────────────────────────────────────

for key in ["cache_e1", "cache_e2", "cache_e3", "cache_e4", "pdf_bytes", "pdf_name",
            "status_e1", "status_e2", "status_e3", "status_e4"]:
    if key not in st.session_state:
        st.session_state[key] = None


# ─────────────────────────────────────────────
# Helper: buat file Excel dari cache 3 engine
# ─────────────────────────────────────────────

def build_excel(daftar_isi: str, keywords: str, ringkasan: str, filename: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Training Data"

    # Styling
    header_font   = Font(bold=True, color="FFFFFF", size=12)
    header_fill   = PatternFill("solid", start_color="0F3460")
    center_align  = Alignment(horizontal="center", vertical="top", wrap_text=True)
    wrap_align    = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Header row
    headers = ["Daftar Isi (Engine 1)", "Keywords (Engine 2)", "Ringkasan (Engine 3)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Baris metadata
    meta_fill = PatternFill("solid", start_color="1E3A5F")
    meta_font = Font(color="A8B2D8", italic=True, size=10)
    meta_label = f"Sumber: {filename}"
    for col_idx in range(1, 4):
        cell = ws.cell(row=2, column=col_idx, value=meta_label if col_idx == 1 else "")
        cell.font   = meta_font
        cell.fill   = meta_fill
        cell.alignment = wrap_align
        cell.border = thin_border

    # Data row
    data_fill = PatternFill("solid", start_color="1E1E2E")
    data_font = Font(color="CDD6F4", size=11)
    data = [daftar_isi or "", keywords or "", ringkasan or ""]
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.font   = data_font
        cell.fill   = data_fill
        cell.alignment = wrap_align
        cell.border = thin_border

    # Lebar kolom
    for col_idx, width in enumerate([50, 40, 80], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Tinggi baris data
    ws.row_dimensions[3].height = 300

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────

st.markdown("""
<div class="main-header">
    <h1>🤖 Training Dataset Generator</h1>
    <p>Upload dokumen PDF → Proses otomatis → Download Excel & JSONL untuk training AI</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SIDEBAR — Upload & Info Engine
# ─────────────────────────────────────────────

with st.sidebar:
    st.markdown("### 📂 Upload Dokumen")
    uploaded = st.file_uploader("Pilih file PDF", type=["pdf"], label_visibility="collapsed")

    if uploaded:
        st.session_state.pdf_bytes = uploaded.read()
        st.session_state.pdf_name  = uploaded.name
        st.success(f"✅ {uploaded.name}")
        st.caption(f"Ukuran: {len(st.session_state.pdf_bytes) / 1024:.1f} KB")

    st.divider()
    st.markdown("### ⚙️ Deskripsi Engine")

    engines_info = [
        ("🔍 Engine 1", "Identifikasi Daftar Isi",
         "Mengekstrak daftar isi, membersihkan titik-titik dan nomor halaman, dipisahkan (;)"),
        ("🏷️ Engine 2", "Ekstraksi Keywords",
         "Menghasilkan kata kunci penting dari isi dokumen dalam Bahasa Indonesia"),
        ("📝 Engine 3", "Ringkasan Dokumen",
         "Meringkas isi dokumen dalam B. Indonesia. Dok. dwibahasa → hanya bagian Indonesia"),
        ("🔄 Engine 4", "Paraphrase → JSONL",
         "Mengidentifikasi, mem-paraphrase, dan mengubah dokumen ke format JSONL training"),
    ]

    for icon_name, title, desc in engines_info:
        st.markdown(f"""
        <div class="engine-card">
            <h4>{icon_name} — {title}</h4>
            <p>{desc}</p>
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# MAIN CONTENT
# ─────────────────────────────────────────────

if not st.session_state.pdf_bytes:
    st.info("⬆️ Upload file PDF di sidebar untuk memulai pemrosesan.")
    st.stop()

col_btn1, col_btn2 = st.columns(2)

with col_btn1:
    run_excel = st.button("▶ Jalankan Engine 1 + 2 + 3 (Excel)", use_container_width=True)

with col_btn2:
    run_jsonl = st.button("▶ Jalankan Engine 4 (JSONL)", use_container_width=True)

run_all = st.button("🚀 Jalankan SEMUA Engine (Excel + JSONL)", use_container_width=True)

st.divider()

# ─── Eksekusi Engine 1, 2, 3 ───
if run_excel or run_all:
    pdf   = st.session_state.pdf_bytes
    fname = st.session_state.pdf_name

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("#### 🔍 Engine 1 — Daftar Isi")
        with st.spinner("Mengidentifikasi daftar isi..."):
            try:
                result = engine1_daftar_isi(pdf)
                st.session_state.cache_e1 = result
                st.session_state.status_e1 = "success"
                st.markdown('<div class="status-box status-success">✅ Selesai</div>', unsafe_allow_html=True)
            except Exception as e:
                st.session_state.status_e1 = "error"
                st.markdown(f'<div class="status-box status-error">❌ Error: {e}</div>', unsafe_allow_html=True)

    with col2:
        st.markdown("#### 🏷️ Engine 2 — Keywords")
        with st.spinner("Mengekstrak keywords..."):
            try:
                result = engine2_keywords(pdf)
                st.session_state.cache_e2 = result
                st.session_state.status_e2 = "success"
                st.markdown('<div class="status-box status-success">✅ Selesai</div>', unsafe_allow_html=True)
            except Exception as e:
                st.session_state.status_e2 = "error"
                st.markdown(f'<div class="status-box status-error">❌ Error: {e}</div>', unsafe_allow_html=True)

    with col3:
        st.markdown("#### 📝 Engine 3 — Ringkasan")
        with st.spinner("Membuat ringkasan..."):
            try:
                result = engine3_ringkasan(pdf)
                st.session_state.cache_e3 = result
                st.session_state.status_e3 = "success"
                st.markdown('<div class="status-box status-success">✅ Selesai</div>', unsafe_allow_html=True)
            except Exception as e:
                st.session_state.status_e3 = "error"
                st.markdown(f'<div class="status-box status-error">❌ Error: {e}</div>', unsafe_allow_html=True)

# ─── Eksekusi Engine 4 ───
if run_jsonl or run_all:
    pdf   = st.session_state.pdf_bytes
    fname = st.session_state.pdf_name

    st.markdown("#### 🔄 Engine 4 — Paraphrase & JSONL")
    with st.spinner("Mem-paraphrase dan mengkonversi ke JSONL... (mungkin butuh beberapa menit)"):
        try:
            result = engine4_jsonl(pdf, fname)
            st.session_state.cache_e4 = result
            st.session_state.status_e4 = "success"
            st.markdown('<div class="status-box status-success">✅ JSONL berhasil dibuat</div>', unsafe_allow_html=True)
        except Exception as e:
            st.session_state.status_e4 = "error"
            st.markdown(f'<div class="status-box status-error">❌ Error: {e}</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# TAMPILKAN HASIL CACHE
# ─────────────────────────────────────────────

if any([st.session_state.cache_e1, st.session_state.cache_e2,
        st.session_state.cache_e3, st.session_state.cache_e4]):

    st.markdown("## 📊 Hasil Pemrosesan")

    tab1, tab2, tab3, tab4 = st.tabs(
        ["📋 Daftar Isi", "🏷️ Keywords", "📝 Ringkasan", "📄 JSONL Preview"]
    )

    with tab1:
        if st.session_state.cache_e1:
            st.markdown("**Daftar Isi (Engine 1)**")
            items = [i.strip() for i in st.session_state.cache_e1.split(";") if i.strip()]
            for i, item in enumerate(items, 1):
                st.markdown(f"**{i}.** {item}")
            st.caption(f"Total: {len(items)} item")
        else:
            st.info("Belum diproses. Jalankan Engine 1.")

    with tab2:
        if st.session_state.cache_e2:
            st.markdown("**Keywords (Engine 2)**")
            keywords = [k.strip() for k in st.session_state.cache_e2.split(",") if k.strip()]
            cols = st.columns(3)
            for i, kw in enumerate(keywords):
                cols[i % 3].markdown(f"🔹 {kw}")
            st.caption(f"Total: {len(keywords)} keyword")
        else:
            st.info("Belum diproses. Jalankan Engine 2.")

    with tab3:
        if st.session_state.cache_e3:
            st.markdown("**Ringkasan (Engine 3)**")
            st.markdown(f"""<div class="result-container">{st.session_state.cache_e3}</div>""",
                        unsafe_allow_html=True)
        else:
            st.info("Belum diproses. Jalankan Engine 3.")

    with tab4:
        if st.session_state.cache_e4:
            st.markdown("**JSONL Preview (Engine 4)**")
            lines = st.session_state.cache_e4.strip().splitlines()
            st.caption(f"Total: {len(lines)} baris JSONL")
            for i, line in enumerate(lines[:10], 1):
                try:
                    obj = json.loads(line)
                    with st.expander(f"Baris {i}: {obj.get('prompt', '')[:60]}..."):
                        st.json(obj)
                except Exception:
                    st.code(line, language="json")
            if len(lines) > 10:
                st.caption(f"... dan {len(lines) - 10} baris lainnya (lihat file download)")
        else:
            st.info("Belum diproses. Jalankan Engine 4.")


# ─────────────────────────────────────────────
# DOWNLOAD SECTION
# ─────────────────────────────────────────────

st.markdown("---")
st.markdown("## ⬇️ Download Hasil")

dl_col1, dl_col2 = st.columns(2)

with dl_col1:
    st.markdown("### 📊 Download Excel")
    st.caption("Berisi: Daftar Isi (Kol A) · Keywords (Kol B) · Ringkasan (Kol C)")

    if st.session_state.cache_e1 or st.session_state.cache_e2 or st.session_state.cache_e3:
        excel_bytes = build_excel(
            st.session_state.cache_e1 or "",
            st.session_state.cache_e2 or "",
            st.session_state.cache_e3 or "",
            st.session_state.pdf_name or "dokumen.pdf"
        )
        base_name = (st.session_state.pdf_name or "dokumen").replace(".pdf", "")
        st.download_button(
            label="⬇️ Download Excel (.xlsx)",
            data=excel_bytes,
            file_name=f"{base_name}_training_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.warning("⚠️ Jalankan Engine 1, 2, atau 3 terlebih dahulu.")

with dl_col2:
    st.markdown("### 📄 Download JSONL")
    st.caption("Format training dataset untuk fine-tuning model AI")

    if st.session_state.cache_e4:
        base_name = (st.session_state.pdf_name or "dokumen").replace(".pdf", "")
        st.download_button(
            label="⬇️ Download JSONL (.jsonl)",
            data=st.session_state.cache_e4.encode("utf-8"),
            file_name=f"{base_name}_training_data.jsonl",
            mime="application/json",
            use_container_width=True,
        )
    else:
        st.warning("⚠️ Jalankan Engine 4 terlebih dahulu.")


# ─────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────

st.markdown("---")
st.markdown(
    "<center><small>🤖 Training Dataset Generator · Powered by Claude AI · "
    "Semua pemrosesan dilakukan secara lokal via API Anthropic</small></center>",
    unsafe_allow_html=True
)
